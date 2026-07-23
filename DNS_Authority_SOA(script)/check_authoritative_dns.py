import dns.resolver
import dns.message
import dns.query
import dns.flags
import dns.rdatatype
import dns.exception
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill


INPUT_FILE = "domains.txt"
OUTPUT_FILE = "dns_authority_soa_report.xlsx"

# Report layout. This list alone controls column order; any key not
# listed here is dropped instead of becoming a stray extra column.
COLUMNS = [
    "Domain",
    "Nameserver",
    "NS_IP",
    "Authoritative",
    "Status",
    "Primary",
    "SOA Status",
]

# Queried in order. A SERVFAIL or timeout from one is retried on the next,
# so a single flaky resolver no longer collapses a domain to one ERROR row.
RESOLVER_IPS = ["8.8.8.8", "1.1.1.1", "9.9.9.9"]

TIMEOUT = 5


def make_resolver(ip):
    r = dns.resolver.Resolver(configure=False)
    r.nameservers = [ip]
    r.timeout = TIMEOUT
    r.lifetime = TIMEOUT
    return r


RESOLVERS = [make_resolver(ip) for ip in RESOLVER_IPS]


def resolve_any(name, rdtype):
    """
    Try every resolver in turn.

    Returns (answers, error). NXDOMAIN and NoAnswer are definitive answers
    about the zone, so they stop the loop; SERVFAIL and timeouts are
    treated as transport failures and retried elsewhere.
    """
    last_error = None

    for r in RESOLVERS:
        try:
            return r.resolve(name, rdtype), None
        except (dns.resolver.NXDOMAIN, dns.resolver.NoAnswer) as exc:
            return None, exc
        except Exception as exc:
            last_error = exc

    return None, last_error


def classify(error):
    if isinstance(error, dns.resolver.NXDOMAIN):
        return "NXDOMAIN"
    if isinstance(error, dns.resolver.NoAnswer):
        return "NO_SOA"
    if isinstance(error, dns.resolver.NoNameservers):
        return "SERVFAIL"
    if isinstance(error, dns.exception.Timeout):
        return "TIMEOUT"
    return "ERROR"


def get_soa_info(domain):
    """Returns (SOA Status, Primary NS, Serial)."""
    answers, error = resolve_any(domain, "SOA")

    if answers is None:
        return classify(error), "", ""

    for rdata in answers:
        return (
            "Published",
            str(rdata.mname).rstrip("."),
            str(rdata.serial),
        )

    return "UNKNOWN", "", ""


def get_nameserver_ip(ns_name):
    answers, _ = resolve_any(ns_name, "A")
    if answers is None:
        return ""
    return answers[0].to_text()


def get_delegation_ns(domain):
    """
    Ask the PARENT zone's nameservers what they delegate this domain to.

    A zone whose own servers are SERVFAILing still has a delegation sitting
    in its parent (com.my, .my, .hk ...), and that delegation is what the
    rest of the internet follows. This is how you get a nameserver list for
    a broken zone.
    """
    labels = domain.split(".")

    for i in range(1, len(labels)):
        parent = ".".join(labels[i:])

        parent_ns, _ = resolve_any(parent, "NS")
        if parent_ns is None:
            continue

        for parent_ns_record in parent_ns:
            parent_ip = get_nameserver_ip(str(parent_ns_record).rstrip("."))
            if not parent_ip:
                continue

            try:
                query = dns.message.make_query(domain, "NS")
                response = dns.query.udp(query, parent_ip, timeout=TIMEOUT)
            except Exception:
                continue

            names = set()
            for rrset in list(response.answer) + list(response.authority):
                if rrset.rdtype == dns.rdatatype.NS:
                    for record in rrset:
                        names.add(str(record.target).rstrip("."))

            if names:
                print(f"  NS recovered from parent zone '{parent}'")
                return sorted(names)

    return []


def get_ns_list(domain, primary_ns):
    """Recursive lookup, then parent delegation, then the SOA's own MNAME."""
    answers, error = resolve_any(domain, "NS")

    if answers is not None:
        return sorted(str(ns).rstrip(".") for ns in answers), classify(None)

    print(f"  NS lookup failed ({classify(error)}), trying parent delegation")

    delegated = get_delegation_ns(domain)
    if delegated:
        return delegated, classify(error)

    if primary_ns:
        print(f"  Falling back to SOA MNAME: {primary_ns}")
        return [primary_ns], classify(error)

    return [], classify(error)


def check_authoritative(domain, ns_ip):
    """UDP first, then TCP — a dropped UDP reply is not a 'No'."""
    query = dns.message.make_query(domain, "SOA")

    try:
        response = dns.query.udp(query, ns_ip, timeout=TIMEOUT)
        return bool(response.flags & dns.flags.AA)
    except dns.exception.Timeout:
        pass
    except Exception:
        return False

    try:
        response = dns.query.tcp(query, ns_ip, timeout=TIMEOUT)
        return bool(response.flags & dns.flags.AA)
    except Exception:
        return False


# Read domains
with open(INPUT_FILE, "r", encoding="utf-8-sig") as f:
    domains = [
        line.strip()
        for line in f
        if line.strip() and not line.strip().startswith("#")
    ]

results = []

for domain in domains:

    print(f"Checking {domain}")

    soa_status, primary_ns, serial = get_soa_info(domain)

    ns_list, ns_error = get_ns_list(domain, primary_ns)

    if not ns_list:
        results.append({
            "Domain": domain,
            "Nameserver": "",
            "NS_IP": "",
            "Authoritative": "",
            "Status": f"NS_{ns_error}",
            "Primary": "",
            "SOA Status": soa_status,
        })
        continue

    domain_rows = []

    for ns_name in ns_list:

        ns_ip = get_nameserver_ip(ns_name)

        authoritative = False
        if ns_ip:
            authoritative = check_authoritative(domain, ns_ip)

        if not ns_ip:
            status = "NO_IP"
        elif authoritative:
            status = "OK"
        else:
            status = "WARNING"

        domain_rows.append({
            "Domain": domain,
            "Nameserver": ns_name,
            "NS_IP": ns_ip,
            "Authoritative": "Yes" if authoritative else "No",
            "Status": status,
            "Primary": "*" if ns_name == primary_ns else "-",
            "SOA Status": "",
        })

    # Show SOA Status only once (middle row if possible)
    if domain_rows:
        middle_row = len(domain_rows) // 2
        domain_rows[middle_row]["SOA Status"] = soa_status

    results.extend(domain_rows)

# Export to Excel
df = pd.DataFrame(results, columns=COLUMNS)

with pd.ExcelWriter(
    OUTPUT_FILE,
    engine="openpyxl"
) as writer:
    df.to_excel(
        writer,
        sheet_name="DNS Audit",
        index=False
    )

# Formatting
wb = load_workbook(OUTPUT_FILE)
ws = wb["DNS Audit"]

header_fill = PatternFill(
    fill_type="solid",
    start_color="D9D9D9",
    end_color="D9D9D9"
)

header_font = Font(bold=True)

# Header formatting
for cell in ws[1]:
    cell.font = header_font
    cell.fill = header_fill

# Auto-size columns
for column in ws.columns:

    max_length = 0
    column_letter = column[0].column_letter

    for cell in column:
        try:
            value = str(cell.value) if cell.value else ""
            if len(value) > max_length:
                max_length = len(value)
        except Exception:
            pass

    ws.column_dimensions[column_letter].width = max_length + 3

# Freeze header row
ws.freeze_panes = "A2"

# Enable filter
ws.auto_filter.ref = ws.dimensions

wb.save(OUTPUT_FILE)

print(f"\nCompleted. Report saved to: {OUTPUT_FILE}")
